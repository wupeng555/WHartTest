from rest_framework import viewsets, permissions, status, filters
from django_filters.rest_framework import DjangoFilterBackend # 导入 DjangoFilterBackend
from rest_framework.response import Response
from rest_framework.decorators import action
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.http import HttpResponse
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io

from .models import (
    TestCase, TestCaseModule, Project, TestCaseScreenshot,
    TestSuite, TestExecution, TestCaseResult
)
from .serializers import TestCaseSerializer, TestCaseModuleSerializer, TestCaseScreenshotSerializer
from .permissions import IsProjectMemberForTestCase, IsProjectMemberForTestCaseModule
from .filters import TestCaseFilter # 导入自定义过滤器
# 确保导入项目自定义的权限类
from wharttest_django.permissions import HasModelPermission, permission_required


def _normalize_media_url(url: str) -> str:
    """
    规范化媒体URL，确保正确添加MEDIA_URL前缀
    避免双重前缀问题（如 /media//media/...）
    """
    if not url:
        return url
    
    # 如果已经是完整的HTTP URL，直接返回
    if url.startswith('http://') or url.startswith('https://'):
        return url
    
    # 规范化路径分隔符（将反斜杠替换为正斜杠）
    url = url.replace('\\', '/')
    
    media_url = settings.MEDIA_URL.rstrip('/')  # 通常是 '/media'
    
    # 如果已经以 MEDIA_URL 开头，直接返回
    if url.startswith(media_url + '/') or url.startswith(media_url):
        return url
    
    # 如果以 / 开头，去掉开头的 /
    if url.startswith('/'):
        url = url[1:]
    
    return f"{media_url}/{url}"

class TestCaseViewSet(viewsets.ModelViewSet):
    """
    用例视图集，处理用例的 CRUD 操作，并支持嵌套创建/更新用例步骤。
    API 端点将嵌套在项目下，例如 /api/projects/{project_pk}/testcases/
    """
    serializer_class = TestCaseSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter] # 添加 DjangoFilterBackend
    filterset_class = TestCaseFilter # 使用自定义的 FilterSet
    search_fields = ['name', 'precondition']

    def get_permissions(self):
        """
        返回此视图所需权限的实例列表。
        这将覆盖 settings.DEFAULT_PERMISSION_CLASSES。
        """
        # 确保所有权限类都被实例化
        return [
            permissions.IsAuthenticated(),
            HasModelPermission(), # 使用支持 @permission_required 装饰器的权限类
            IsProjectMemberForTestCase()
        ]

    def get_queryset(self):
        """
        根据 URL 中的 project_pk 过滤用例。
        确保只返回指定项目下的用例。
        """
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            project = get_object_or_404(Project, pk=project_pk)
            # 权限类 IsProjectMemberForTestCase 已经检查了用户是否是此项目的成员
            # 所以这里可以直接返回项目下的用例
            return TestCase.objects.filter(project=project).select_related('creator', 'module').prefetch_related('steps')
        # 如果没有 project_pk (理论上不应该发生，因为路由是嵌套的)
        # 返回空 queryset 或根据需求抛出错误
        return TestCase.objects.none()

    def perform_create(self, serializer):
        """
        在创建用例时，自动关联项目和创建人。
        """
        project_pk = self.kwargs.get('project_pk')
        project = get_object_or_404(Project, pk=project_pk)
        # 权限类 IsProjectMemberForTestCase 已经确保用户是项目成员
        serializer.save(creator=self.request.user, project=project)

    # create 和 update 方法将使用序列化器中定义的嵌套写入逻辑。
    # DRF 的 ModelViewSet 会自动调用 serializer.save()，
    # 其中包含了处理嵌套 'steps' 的逻辑。

    # 如果需要更细致的控制，可以覆盖 create 和 update 方法，例如：
    # def create(self, request, *args, **kwargs):
    #     project_pk = self.kwargs.get('project_pk')
    #     project = get_object_or_404(Project, pk=project_pk)
    #
    #     # 可以在这里添加额外的逻辑，例如检查项目状态等
    #
    #     serializer = self.get_serializer(data=request.data)
    #     serializer.is_valid(raise_exception=True)
    #     self.perform_create(serializer) # perform_create 中会设置 project 和 creator
    #     headers = self.get_success_headers(serializer.data)
    #     return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object() # get_object 会进行对象级权限检查

        # 可以在这里添加额外的逻辑

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        
        self.perform_update(serializer) # perform_update 默认只调用 serializer.save()

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)

    # perform_update 默认调用 serializer.save()，我们的序列化器 update 方法会处理嵌套步骤。
    # perform_destroy 默认调用 instance.delete()。

    @action(detail=False, methods=['get', 'post'], url_path='export-excel')
    def export_excel(self, request, project_pk=None):
        """
        导出用例为Excel格式
        支持两种方式传递要导出的用例ID：
        1. GET请求通过ids参数: /api/projects/1/testcases/export-excel/?ids=1,2,3
        2. POST请求通过请求体: {"ids": [1, 2, 3]}
        如果不提供ids，则导出项目下所有用例
        """
        testcase_ids = None

        if request.method == 'POST':
            # POST请求，从请求体获取ids
            ids_data = request.data.get('ids', [])
            if ids_data:
                try:
                    testcase_ids = [int(id) for id in ids_data]
                except (ValueError, TypeError):
                    from rest_framework.response import Response
                    return Response(
                        {'error': 'ids参数格式错误，应为数字列表'},
                        status=400
                    )
        else:
            # GET请求，从查询参数获取ids
            ids_param = request.query_params.get('ids', '')
            if ids_param:
                try:
                    testcase_ids = [int(id.strip()) for id in ids_param.split(',') if id.strip()]
                except ValueError:
                    from rest_framework.response import Response
                    return Response(
                        {'error': 'ids参数格式错误，应为逗号分隔的数字列表'},
                        status=400
                    )

        # 根据是否提供了ids来过滤queryset
        if testcase_ids:
            queryset = self.get_queryset().filter(id__in=testcase_ids)
        else:
            queryset = self.get_queryset()

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        # 设置表头
        headers = [
            '用例名称', '所属模块', '标签', '前置条件',
            '步骤描述', '预期结果', '编辑模式', '备注', '用例等级'
        ]

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 写入数据
        for row, testcase in enumerate(queryset, 2):
            # 获取模块路径
            module_path = self._get_module_path(testcase.module) if testcase.module else ""

            # 获取步骤描述和预期结果
            steps_desc, expected_results = self._format_steps(testcase.steps.all())

            # 写入数据行
            ws.cell(row=row, column=1, value=testcase.name)
            ws.cell(row=row, column=2, value=module_path)
            ws.cell(row=row, column=3, value="")  # 标签字段，当前数据库中没有
            ws.cell(row=row, column=4, value=testcase.precondition or "")
            ws.cell(row=row, column=5, value=steps_desc)
            ws.cell(row=row, column=6, value=expected_results)
            ws.cell(row=row, column=7, value="STEP")  # 编辑模式，固定为STEP
            ws.cell(row=row, column=8, value=testcase.notes or "")
            ws.cell(row=row, column=9, value=testcase.level)

        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 创建HTTP响应
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # 获取项目名称用于文件名
        project = get_object_or_404(Project, pk=project_pk)
        filename = f"{project.name}_测试用例.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def _get_module_path(self, module):
        """
        获取模块的完整路径
        """
        if not module:
            return ""

        path_parts = []
        current = module
        while current:
            path_parts.insert(0, current.name)
            current = current.parent

        return "/" + "/".join(path_parts)

    def _format_steps(self, steps):
        """
        格式化步骤描述和预期结果
        """
        steps_desc = []
        expected_results = []

        for step in steps.order_by('step_number'):
            steps_desc.append(f"[{step.step_number}]{step.description}")
            expected_results.append(f"[{step.step_number}]{step.expected_result}")

        return "\n".join(steps_desc), "\n".join(expected_results)

    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request, **kwargs):
        """
        批量删除用例
        POST请求体格式: {"ids": [1, 2, 3, 4]}
        """
        # 获取要删除的用例ID列表
        ids_data = request.data.get('ids', [])

        if not ids_data:
            return Response(
                {'error': '请提供要删除的用例ID列表'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 验证ID格式
        try:
            testcase_ids = [int(id) for id in ids_data]
        except (ValueError, TypeError):
            return Response(
                {'error': 'ids参数格式错误，应为数字列表'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not testcase_ids:
            return Response(
                {'error': '用例ID列表不能为空'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 获取当前项目下的用例queryset，确保数据隔离
        queryset = self.get_queryset()

        # 过滤出要删除的用例，确保只能删除当前项目下的用例
        testcases_to_delete = queryset.filter(id__in=testcase_ids)

        # 检查是否所有请求的ID都存在
        found_ids = list(testcases_to_delete.values_list('id', flat=True))
        not_found_ids = [id for id in testcase_ids if id not in found_ids]

        if not_found_ids:
            return Response(
                {
                    'error': f'以下用例ID不存在或不属于当前项目: {not_found_ids}',
                    'not_found_ids': not_found_ids
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # 记录删除前的信息用于返回
        deleted_testcases_info = []
        for testcase in testcases_to_delete:
            deleted_testcases_info.append({
                'id': testcase.id,
                'name': testcase.name,
                'module': testcase.module.name if testcase.module else None
            })

        # 执行批量删除
        try:
            with transaction.atomic():
                # 删除用例（关联的步骤会因为外键级联删除而自动删除）
                deleted_count, deleted_details = testcases_to_delete.delete()

                return Response({
                    'message': f'成功删除 {len(deleted_testcases_info)} 个用例',
                    'deleted_count': len(deleted_testcases_info),
                    'deleted_testcases': deleted_testcases_info,
                    'deletion_details': deleted_details
                }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': f'删除过程中发生错误: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'], url_path='upload-screenshots')
    @permission_required('testcases.add_testcasescreenshot')
    def upload_screenshots(self, request, project_pk=None, pk=None):
        """
        上传测试用例截屏（支持多张图片）
        POST /api/projects/{project_pk}/testcases/{pk}/upload-screenshots/
        请求体: multipart/form-data
        支持字段:
        - screenshots: 图片文件（可多个）
        - title: 图片标题（可选）
        - description: 图片描述（可选）
        - step_number: 对应步骤编号（可选）
        - mcp_session_id: MCP会话ID（可选）
        - page_url: 页面URL（可选）
        """
        testcase = self.get_object()

        # 获取上传的文件
        uploaded_files = request.FILES.getlist('screenshots')
        if not uploaded_files:
            # 兼容单文件上传
            if 'screenshot' in request.FILES:
                uploaded_files = [request.FILES['screenshot']]
            else:
                return Response(
                    {'error': '请提供截屏文件，字段名为 screenshots 或 screenshot'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # 验证文件数量限制
        if len(uploaded_files) > 10:
            return Response(
                {'error': '一次最多只能上传10张图片'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 验证文件类型和大小
        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif']
        max_size = 5 * 1024 * 1024  # 5MB

        for file in uploaded_files:
            if file.content_type not in allowed_types:
                return Response(
                    {'error': f'文件 {file.name} 格式不支持，只支持 JPEG、PNG、GIF 格式'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if file.size > max_size:
                return Response(
                    {'error': f'文件 {file.name} 大小超过5MB限制'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        try:
            created_screenshots = []

            # 获取额外信息
            title = request.data.get('title', '')
            description = request.data.get('description', '')
            step_number = request.data.get('step_number')
            mcp_session_id = request.data.get('mcp_session_id', '')
            page_url = request.data.get('page_url', '')

            # 处理step_number
            if step_number:
                try:
                    step_number = int(step_number)
                except (ValueError, TypeError):
                    step_number = None

            # 为每个文件创建截屏记录
            for i, file in enumerate(uploaded_files):
                screenshot_data = {
                    'test_case': testcase.id,
                    'screenshot': file,
                    'title': f"{title} ({i+1})" if title and len(uploaded_files) > 1 else title,
                    'description': description,
                    'step_number': step_number,
                    'mcp_session_id': mcp_session_id,
                    'page_url': page_url,
                }

                serializer = TestCaseScreenshotSerializer(
                    data=screenshot_data,
                    context={'request': request}
                )

                if serializer.is_valid():
                    screenshot = serializer.save()
                    created_screenshots.append(serializer.data)
                else:
                    return Response(
                        {'error': f'文件 {file.name} 保存失败: {serializer.errors}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            return Response({
                'message': f'成功上传 {len(created_screenshots)} 张截屏',
                'screenshots': created_screenshots
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': f'上传失败: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], url_path='screenshots')
    def list_screenshots(self, request, project_pk=None, pk=None):
        """
        获取测试用例的所有截屏
        GET /api/projects/{project_pk}/testcases/{pk}/screenshots/
        """
        testcase = self.get_object()
        screenshots = testcase.screenshots.all()
        serializer = TestCaseScreenshotSerializer(
            screenshots,
            many=True,
            context={'request': request}
        )
        return Response(serializer.data)

    @action(detail=True, methods=['delete'], url_path='screenshots/(?P<screenshot_id>[^/.]+)')
    @permission_required('testcases.delete_testcasescreenshot')
    def delete_screenshot(self, request, project_pk=None, pk=None, screenshot_id=None):
        """
        删除指定的截屏
        DELETE /api/projects/{project_pk}/testcases/{pk}/screenshots/{screenshot_id}/
        """
        testcase = self.get_object()

        try:
            screenshot = testcase.screenshots.get(id=screenshot_id)
            screenshot.delete()
            return Response({
                'message': '截屏删除成功'
            }, status=status.HTTP_200_OK)
        except TestCaseScreenshot.DoesNotExist:
            return Response(
                {'error': '截屏不存在'},
                status=status.HTTP_404_NOT_FOUND
            )
    @action(detail=True, methods=['post'], url_path='screenshots/batch-delete')
    @permission_required('testcases.delete_testcasescreenshot')
    def batch_delete_screenshots(self, request, project_pk=None, pk=None):
        """
        批量删除测试用例的截屏
        POST /api/projects/{project_pk}/testcases/{pk}/screenshots/batch-delete/
        请求体: {"ids": [1, 2, 3]}
        """
        testcase = self.get_object()
        
        # 获取要删除的截图ID列表
        ids_data = request.data.get('ids', [])
        
        if not ids_data:
            return Response(
                {'error': '请提供要删除的截图ID列表'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 验证ID格式
        try:
            screenshot_ids = [int(id) for id in ids_data]
        except (ValueError, TypeError):
            return Response(
                {'error': 'ids参数格式错误，应为数字列表'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not screenshot_ids:
            return Response(
                {'error': '截图ID列表不能为空'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 过滤出要删除的截图，确保只能删除当前测试用例下的截图
        screenshots_to_delete = testcase.screenshots.filter(id__in=screenshot_ids)
        
        # 检查是否所有请求的ID都存在
        found_ids = list(screenshots_to_delete.values_list('id', flat=True))
        not_found_ids = [id for id in screenshot_ids if id not in found_ids]
        
        if not_found_ids:
            return Response(
                {
                    'error': f'以下截图ID不存在或不属于当前测试用例: {not_found_ids}',
                    'not_found_ids': not_found_ids
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 记录删除前的信息
        deleted_screenshots_info = []
        for screenshot in screenshots_to_delete:
            deleted_screenshots_info.append({
                'id': screenshot.id,
                'title': screenshot.title or '无标题',
                'step_number': screenshot.step_number
            })
        
        # 执行批量删除
        try:
            with transaction.atomic():
                deleted_count, _ = screenshots_to_delete.delete()
                
                return Response({
                    'message': f'成功删除 {len(deleted_screenshots_info)} 张截图',
                    'deleted_count': len(deleted_screenshots_info),
                    'deleted_screenshots': deleted_screenshots_info
                }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response(
                {'error': f'删除过程中发生错误: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



class TestCaseModuleViewSet(viewsets.ModelViewSet):
    """
    用例模块视图集，处理模块的 CRUD 操作，支持5级子模块。
    API 端点将嵌套在项目下，例如 /api/projects/{project_pk}/testcase-modules/
    """
    serializer_class = TestCaseModuleSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']

    def get_permissions(self):
        """
        返回此视图所需权限的实例列表。
        """
        return [
            permissions.IsAuthenticated(),
            HasModelPermission(),
            IsProjectMemberForTestCaseModule()
        ]

    def get_queryset(self):
        """
        根据 URL 中的 project_pk 过滤模块。
        确保只返回指定项目下的模块。
        """
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            project = get_object_or_404(Project, pk=project_pk)
            # 权限类 IsProjectMemberForTestCaseModule 已经检查了用户是否是此项目的成员
            return TestCaseModule.objects.filter(project=project).select_related('creator', 'parent')
        return TestCaseModule.objects.none()

    def perform_create(self, serializer):
        """
        在创建模块时，自动关联项目和创建人。
        """
        project_pk = self.kwargs.get('project_pk')
        project = get_object_or_404(Project, pk=project_pk)
        # 将项目实例添加到序列化器上下文，用于验证
        serializer.context['project'] = project
        # 保存模块，设置创建人和项目
        serializer.save(creator=self.request.user, project=project)

    def perform_destroy(self, instance):
        """
        删除模块前检查是否有关联的测试用例
        """
        if instance.testcases.exists():
            from rest_framework.exceptions import ValidationError
            testcase_count = instance.testcases.count()
            raise ValidationError(
                f"无法删除模块 '{instance.name}'，因为该模块下还有 {testcase_count} 个测试用例。请先删除或移动这些用例。"
            )
        instance.delete()

    def get_serializer_context(self):
        """
        为序列化器提供额外的上下文。
        """
        context = super().get_serializer_context()
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            project = get_object_or_404(Project, pk=project_pk)
            context['project'] = project
        return context


class TestSuiteViewSet(viewsets.ModelViewSet):
    """
    测试套件视图集，处理测试套件的 CRUD 操作
    API 端点将嵌套在项目下，例如 /api/projects/{project_pk}/test-suites/
    """
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'description']

    def get_permissions(self):
        """返回此视图所需权限的实例列表"""
        from .permissions import IsProjectMemberForTestSuite
        return [
            permissions.IsAuthenticated(),
            HasModelPermission(),
            IsProjectMemberForTestSuite()
        ]

    def get_queryset(self):
        """根据 URL 中的 project_pk 过滤测试套件"""
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            project = get_object_or_404(Project, pk=project_pk)
            return TestSuite.objects.filter(project=project).prefetch_related('testcases', 'creator')
        return TestSuite.objects.none()

    def get_serializer_class(self):
        """根据不同action返回不同的序列化器"""
        from .serializers import TestSuiteSerializer
        return TestSuiteSerializer

    def get_serializer_context(self):
        """为序列化器提供额外的上下文"""
        context = super().get_serializer_context()
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            context['project_id'] = int(project_pk)
        return context
    
    def perform_create(self, serializer):
        """在创建测试套件时，自动关联项目和创建人"""
        project_pk = self.kwargs.get('project_pk')
        project = get_object_or_404(Project, pk=project_pk)
        serializer.save(creator=self.request.user, project=project)


class TestExecutionViewSet(viewsets.ModelViewSet):
    """
    测试执行视图集，处理测试执行的创建、查看和管理
    API 端点将嵌套在项目下，例如 /api/projects/{project_pk}/test-executions/
    """
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['suite__name']
    ordering_fields = ['created_at', 'started_at', 'completed_at', 'status']
    ordering = ['-created_at']

    def get_permissions(self):
        """返回此视图所需权限的实例列表"""
        from .permissions import IsProjectMemberForTestExecution
        return [
            permissions.IsAuthenticated(),
            HasModelPermission(),
            IsProjectMemberForTestExecution()
        ]

    def get_queryset(self):
        """根据 URL 中的 project_pk 过滤测试执行"""
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            project = get_object_or_404(Project, pk=project_pk)
            return TestExecution.objects.filter(
                suite__project=project
            ).select_related('suite', 'executor').prefetch_related('results')
        return TestExecution.objects.none()

    def get_serializer_class(self):
        """根据不同action返回不同的序列化器"""
        from .serializers import TestExecutionSerializer, TestExecutionCreateSerializer
        if self.action == 'create':
            return TestExecutionCreateSerializer
        return TestExecutionSerializer

    def create(self, request, *args, **kwargs):
        """创建测试执行并启动Celery任务"""
        from .serializers import TestExecutionCreateSerializer, TestExecutionSerializer
        from .tasks import execute_test_suite
        
        serializer = TestExecutionCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        suite_id = serializer.validated_data['suite_id']
        suite = get_object_or_404(TestSuite, id=suite_id)
        
        # 验证套件属于当前项目
        project_pk = self.kwargs.get('project_pk')
        if suite.project_id != int(project_pk):
            return Response({
                'error': '测试套件不属于当前项目'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 创建执行记录
        execution = TestExecution.objects.create(
            suite=suite,
            executor=request.user,
            status='pending'
        )
        
        # 使用transaction.on_commit()确保数据库事务提交后再启动Celery任务
        # Django和Celery在同一容器中运行,共享同一数据库连接,避免查询不到记录的问题
        def start_execution_task():
            task = execute_test_suite.delay(execution.id)
            # 更新celery_task_id
            TestExecution.objects.filter(id=execution.id).update(celery_task_id=task.id)
        
        transaction.on_commit(start_execution_task)
        
        # 返回创建的执行记录
        result_serializer = TestExecutionSerializer(execution, context={'request': request})
        return Response(result_serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='cancel')
    def cancel(self, request, project_pk=None, pk=None):
        """取消测试执行"""
        from .tasks import cancel_test_execution
        from celery import current_app
        
        execution = self.get_object()
        
        if execution.status not in ['pending', 'running']:
            return Response({
                'error': f'无法取消状态为 {execution.get_status_display()} 的执行'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 尝试撤销Celery任务
        if execution.celery_task_id:
            current_app.control.revoke(execution.celery_task_id, terminate=True)
        
        # 调用取消任务
        cancel_test_execution.delay(execution.id)
        
        return Response({
            'message': '测试执行取消请求已发送'
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'], url_path='results')
    def results(self, request, project_pk=None, pk=None):
        """获取测试执行的所有结果"""
        from .serializers import TestCaseResultSerializer
        
        execution = self.get_object()
        results = execution.results.all().select_related('testcase')
        serializer = TestCaseResultSerializer(results, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='report')
    def report(self, request, project_pk=None, pk=None):
        """生成测试执行报告"""
        execution = self.get_object()
        
        report_data = {
            'execution_id': execution.id,
            'suite': {
                'id': execution.suite.id,
                'name': execution.suite.name,
                'description': execution.suite.description,
            },
            'executor': {
                'id': execution.executor.id,
                'username': execution.executor.username,
            } if execution.executor else None,
            'status': execution.status,
            'started_at': execution.started_at,
            'completed_at': execution.completed_at,
            'duration': execution.duration,
            'statistics': {
                'total': execution.total_count,
                'passed': execution.passed_count,
                'failed': execution.failed_count,
                'skipped': execution.skipped_count,
                'error': execution.error_count,
                'pass_rate': execution.pass_rate,
            },
            'results': []
        }
        
        # 添加用例执行结果
        for result in execution.results.all().select_related('testcase'):
            screenshots_urls = [
                _normalize_media_url(path) for path in (result.screenshots or [])
            ]
            report_data['results'].append({
                'testcase_id': result.testcase.id,
                'testcase_name': result.testcase.name,
                'status': result.status,
                'error_message': result.error_message,
                'execution_time': result.execution_time,
                'screenshots': screenshots_urls,
            })
        
        # 添加脚本执行结果
        report_data['script_results'] = []
        for script_result in execution.script_results.all().select_related('script'):
            screenshots_urls = [
                _normalize_media_url(path) for path in (script_result.screenshots or [])
            ]
            videos_urls = [
                _normalize_media_url(path) for path in (script_result.videos or [])
            ]
            report_data['script_results'].append({
                'script_id': script_result.script.id,
                'script_name': script_result.script.name,
                'status': script_result.status,
                'error_message': script_result.error_message,
                'execution_time': script_result.execution_time,
                'output': script_result.output,
                'screenshots': screenshots_urls,
                'videos': videos_urls,
            })
        
        return Response(report_data)
    
    def destroy(self, request, *args, **kwargs):
        """
        删除测试执行记录
        只允许删除已完成、失败或已取消的执行记录
        """
        execution = self.get_object()
        
        # 检查执行状态，不允许删除正在运行或等待中的执行
        if execution.status in ['pending', 'running']:
            return Response({
                'error': f'无法删除状态为"{execution.get_status_display()}"的执行记录，请先取消执行'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 记录删除信息用于日志
        execution_info = {
            'id': execution.id,
            'suite_name': execution.suite.name,
            'status': execution.status,
            'created_at': execution.created_at
        }
        
        # 执行删除（关联的TestCaseResult会被级联删除）
        execution.delete()
        
        return Response({
            'message': f'测试执行记录已删除',
            'deleted_execution': execution_info
        }, status=status.HTTP_200_OK)


# ====================== 自动化用例相关 ======================

from .models import AutomationScript, ScriptExecution
from .serializers import (
    AutomationScriptSerializer, AutomationScriptListSerializer,
    ScriptExecutionSerializer, ExecuteScriptSerializer
)
from .script_executor import execute_automation_script


class AutomationScriptViewSet(viewsets.ModelViewSet):
    """
    自动化用例视图集
    
    提供脚本的 CRUD 操作、生成和执行功能
    """
    queryset = AutomationScript.objects.all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description', 'test_case__name']
    ordering_fields = ['created_at', 'updated_at', 'name', 'version']
    ordering = ['-created_at']
    filterset_fields = ['test_case', 'script_type', 'source', 'status', 'test_case__project']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return AutomationScriptListSerializer
        return AutomationScriptSerializer
    
    def get_permissions(self):
        return [
            permissions.IsAuthenticated(),
            HasModelPermission(),
        ]
    
    def get_queryset(self):
        queryset = AutomationScript.objects.select_related(
            'test_case', 'test_case__project', 'creator', 'source_task'
        )
        
        # 支持按项目过滤
        project_id = self.request.query_params.get('project_id')
        if project_id:
            queryset = queryset.filter(test_case__project_id=project_id)
        
        # 非管理员只能看到自己所属项目的脚本
        user = self.request.user
        if not user.is_superuser:
            # 获取用户所属的项目 ID 列表
            user_project_ids = user.project_memberships.values_list('project_id', flat=True)
            queryset = queryset.filter(test_case__project_id__in=user_project_ids)
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(creator=self.request.user)
    
    def _check_project_access(self, project_id):
        """检查用户是否有项目访问权限"""
        user = self.request.user
        if user.is_superuser:
            return True
        return user.project_memberships.filter(project_id=project_id).exists()
    
    @action(detail=True, methods=['post'], url_path='execute')
    def execute(self, request, pk=None):
        """
        执行自动化用例
        
        POST /api/automation-scripts/{id}/execute/
        {
            "headless": true,
            "record_video": false
        }
        """
        script = self.get_object()
        
        # 权限检查：验证用户有该项目的访问权限
        if not self._check_project_access(script.test_case.project_id):
            return Response(
                {'error': '无权访问该项目'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = ExecuteScriptSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        headless = serializer.validated_data.get('headless', script.headless)
        record_video = serializer.validated_data.get('record_video', False)
        
        try:
            execution = execute_automation_script(
                script=script,
                executor=request.user,
                headless=headless,
                record_video=record_video
            )
            
            return Response(
                ScriptExecutionSerializer(execution).data,
                status=status.HTTP_201_CREATED
            )
            
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'error': f'执行脚本失败: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='executions')
    def executions(self, request, pk=None):
        """
        获取脚本的执行历史
        
        GET /api/automation-scripts/{id}/executions/
        """
        script = self.get_object()
        executions = script.executions.all().order_by('-created_at')
        
        # 分页
        page = self.paginate_queryset(executions)
        if page is not None:
            serializer = ScriptExecutionSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = ScriptExecutionSerializer(executions, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], url_path='regenerate')
    def regenerate(self, request, pk=None):
        """
        重新生成脚本内容（基于 recorded_steps）
        
        POST /api/automation-scripts/{id}/regenerate/
        {
            "use_pytest": true
        }
        """
        script = self.get_object()
        
        if not script.recorded_steps:
            return Response(
                {'error': '该脚本没有记录的操作步骤，无法重新生成'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        use_pytest = request.data.get('use_pytest', True)
        
        try:
            from .script_generator import PlaywrightScriptGenerator
            
            generator = PlaywrightScriptGenerator(use_pytest=use_pytest)
            new_content = generator.generate_script(
                recorded_steps=script.recorded_steps,
                test_case_name=script.test_case.name,
                target_url=script.target_url or '',
                timeout_seconds=script.timeout_seconds,
                headless=script.headless,
                description=script.description or ''
            )
            
            # 更新脚本内容和版本
            script.script_content = new_content
            script.version += 1
            script.save()
            
            return Response(AutomationScriptSerializer(script).data)
            
        except Exception as e:
            return Response(
                {'error': f'重新生成失败: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ScriptExecutionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    脚本执行记录视图集（只读）
    """
    queryset = ScriptExecution.objects.all()
    serializer_class = ScriptExecutionSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['script', 'status', 'executor']
    ordering_fields = ['created_at', 'execution_time']
    ordering = ['-created_at']
    
    def get_permissions(self):
        return [
            permissions.IsAuthenticated(),
            HasModelPermission(),
        ]
    
    def get_queryset(self):
        queryset = ScriptExecution.objects.select_related(
            'script', 'script__test_case', 'script__test_case__project', 'executor'
        )
        
        # 非管理员只能看到自己所属项目的执行记录
        user = self.request.user
        if not user.is_superuser:
            user_project_ids = user.project_memberships.values_list('project_id', flat=True)
            queryset = queryset.filter(script__test_case__project_id__in=user_project_ids)
        
        return queryset
