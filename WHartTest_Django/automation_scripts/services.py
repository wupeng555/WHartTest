import os
import json
import yaml
import uuid
import asyncio
import subprocess
import tempfile
import shutil
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from django.conf import settings
from django.utils import timezone
from django.core.files.storage import default_storage
import logging

from .models import AutomationScript, ScriptExecution, ExecutionLog

logger = logging.getLogger(__name__)


class YAMLScriptGenerator:
    """YAML脚本生成服务"""
    
    def __init__(self, api_key: str, api_endpoint: str, model: str = 'qwen-turbo'):
        self.api_key = api_key
        self.api_endpoint = api_endpoint
        self.model = model
    
    def generate_yaml_script(self, test_cases: str, script_type: str, target_url: str = '') -> str:
        """生成YAML脚本"""
        try:
            # 构建提示词
            prompt = self._build_prompt(test_cases, script_type, target_url)
            
            # 调用AI生成脚本
            yaml_content = self._call_ai_api(prompt)
            
            # 验证YAML格式
            self._validate_yaml(yaml_content)
            
            return yaml_content
            
        except Exception as e:
            logger.error(f"生成YAML脚本失败: {str(e)}")
            raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise Exception(f"生成脚本失败: {str(e)}")
    
    def _build_prompt(self, test_cases: str, script_type: str, target_url: str) -> str:
        """构建AI提示词"""
        
        base_prompt = """你是一位专业的自动化测试工程师，请根据提供的测试用例生成Midscene.js YAML自动化脚本。

# Midscene.js YAML 格式说明

## 基本结构
YAML脚本包含环境配置（web/android/ios）、可选的agent配置和tasks任务定义。

## Web环境配置示例：
```yaml
web:
  url: https://www.example.com
  viewportWidth: 1280
  viewportHeight: 960
  waitForNetworkIdle:
    timeout: 5000

tasks:
  - name: 测试任务名称
    flow:
      - ai: 操作描述
      - aiAssert: 验证描述
```

## 支持的操作类型：
1. **ai**: AI智能操作 - 使用自然语言描述操作
2. **aiAssert**: AI智能断言 - 验证页面状态
3. **sleep**: 等待时间（毫秒）
4. **runAdbShell**: 执行ADB命令（Android）
5. **javascript**: 执行JavaScript代码

## 重要规则：
1. 使用中文描述操作和验证
2. 每个task代表一个完整的测试用例
3. flow中的步骤按顺序执行
4. 确保YAML语法正确

请严格按照以上格式生成脚本。"""
        
        if script_type == 'web':
            environment_config = f"""
web:
  url: {target_url or 'https://example.com'}
  viewportWidth: 1280
  viewportHeight: 960
  waitForNetworkIdle:
    timeout: 5000
"""
        elif script_type == 'android':
            environment_config = """
android:
  packageName: com.example.app
  activityName: .MainActivity
"""
        else:  # ios
            environment_config = """
ios:
  bundleId: com.example.app
"""
        
        prompt = f"""{base_prompt}

## 环境配置
{environment_config}

## 测试用例内容
{test_cases}

请生成完整的YAML脚本，包含环境配置和所有测试任务。"""
        
        return prompt
    
    def _call_ai_api(self, prompt: str) -> str:
        """调用AI API"""
        try:
            import requests
            
            headers = {
                'Authorization': f'Bearer {self.api_key}',
                'Content-Type': 'application/json'
            }
            
            data = {
                'model': self.model,
                'messages': [
                    {
                        'role': 'user',
                        'content': prompt
                    }
                ],
                'temperature': 0.3,
                'max_tokens': 4000
            }
            
            response = requests.post(
                f"{self.api_endpoint}/v1/chat/completions",
                headers=headers,
                json=data,
                timeout=60
            )
            
            if response.status_code != 200:
                raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise Exception(f"AI API调用失败: {response.status_code} - {response.text}")
            
            result = response.json()
            content = result['choices'][0]['message']['content']
            
            # 提取YAML内容
            if '```yaml' in content:
                yaml_content = content.split('```yaml')[1].split('```')[0].strip()
            elif '```' in content:
                yaml_content = content.split('```')[1].split('```')[0].strip()
            else:
                yaml_content = content.strip()
            
            return yaml_content
            
        except Exception as e:
            logger.error(f"AI API调用失败: {str(e)}")
            raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise Exception(f"AI服务调用失败: {str(e)}")
    
    def _validate_yaml(self, yaml_content: str) -> None:
        """验证YAML格式"""
        try:
            yaml.safe_load(yaml_content)
        except yaml.YAMLError as e:
            raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise Exception(f"YAML格式错误: {str(e)}")


class MidsceneExecutor:
    """Midscene.js执行器"""
    
    def __init__(self):
        self.base_dir = getattr(settings, 'AUTOMATION_SCRIPTS_DIR', '/tmp/automation_scripts')
        os.makedirs(self.base_dir, exist_ok=True)
    
    async def execute_script(self, script: AutomationScript, executor_user) -> ScriptExecution:
        """执行脚本"""
        try:
            # 创建执行记录
            execution = ScriptExecution.objects.create(
                script=script,
                executor=executor_user,
                execution_id=str(uuid.uuid4())[:8],
                status='pending'
            )
            
            # 创建执行目录
            execution_dir = os.path.join(self.base_dir, execution.execution_id)
            os.makedirs(execution_dir, exist_ok=True)
            
            # 准备执行环境
            await self._prepare_execution_environment(script, execution, execution_dir)
            
            # 执行脚本
            await self._run_midscene_script(execution, execution_dir)
            
            # 处理执行结果
            await self._process_execution_result(execution, execution_dir)
            
            return execution
            
        except Exception as e:
            logger.error(f"脚本执行失败: {str(e)}")
            execution.status = 'failed'
            execution.stderr = str(e)
            execution.completed_at = timezone.now()
            execution.save()
            
            # 记录错误日志
            ExecutionLog.objects.create(
                execution=execution,
                level='error',
                message=f"执行失败: {str(e)}"
            )
            
            raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise
    
    async def _prepare_execution_environment(self, script: AutomationScript, execution: ScriptExecution, execution_dir: str):
        """准备执行环境"""
        try:
            # 保存YAML脚本文件
            yaml_file_path = os.path.join(execution_dir, f"{script.name}.yaml")
            with open(yaml_file_path, 'w', encoding='utf-8') as f:
                f.write(script.yaml_content)
            
            # 创建.env文件
            env_content = self._generate_env_content(script)
            env_file_path = os.path.join(execution_dir, '.env')
            with open(env_file_path, 'w', encoding='utf-8') as f:
                f.write(env_content)
            
            # 创建config.yaml文件
            config_content = self._generate_config_content(script, execution_dir)
            config_file_path = os.path.join(execution_dir, 'config.yaml')
            with open(config_file_path, 'w', encoding='utf-8') as f:
                f.write(config_content)
            
            # 记录准备日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message=f"执行环境准备完成: {execution_dir}"
            )
            
        except Exception as e:
            logger.error(f"准备执行环境失败: {str(e)}")
            raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise Exception(f"环境准备失败: {str(e)}")
    
    def _generate_env_content(self, script: AutomationScript) -> str:
        """生成.env文件内容"""
        return f"""# Midscene AI 配置
MIDSCENE_MODEL_TEXT_ONLY={script.ai_model}
MIDSCENE_MODEL_VISION={script.ai_model}
MIDSCENE_OPENAI_INIT_CONFIG_API_KEY={script.api_key}
MIDSCENE_OPENAI_INIT_CONFIG_BASE_URL={script.api_endpoint}

# 执行配置
MIDSCENE_DEBUG=true
MIDSCENE_SAVE_SCREENSHOT=true
"""
    
    def _generate_config_content(self, script: AutomationScript, execution_dir: str) -> str:
        """生成config.yaml文件内容"""
        config = {
            'tasks': [f"{script.name}.yaml"],
            'output': {
                'reportDir': './midscene_run/report',
                'screenshotDir': './midscene_run/screenshots'
            },
            'execution': {
                'timeout': script.execution_timeout * 1000,  # 转换为毫秒
                'retries': script.retry_count,
                'parallel': False
            }
        }
        
        return yaml.dump(config, default_flow_style=False, allow_unicode=True)
    
    async def _run_midscene_script(self, execution: ScriptExecution, execution_dir: str):
        """运行Midscene脚本"""
        try:
            # 更新执行状态
            execution.status = 'running'
            execution.started_at = timezone.now()
            execution.save()
            
            # 记录开始执行日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message="开始执行Midscene脚本"
            )
            
            # 构建执行命令
            cmd = [
                'npx', 'midscene', 'run',
                '--config', 'config.yaml',
                '--reporter', 'html'
            ]
            
            # 执行命令
            process = await asyncio.create_subprocess_exec(
                *cmd,
                cwd=execution_dir,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
                env={**os.environ, 'NODE_ENV': 'production'}
            )
            
            # 等待执行完成
            stdout, stderr = await process.communicate()
            
            # 更新执行结果
            execution.exit_code = process.returncode
            execution.stdout = stdout.decode('utf-8', errors='ignore')
            execution.stderr = stderr.decode('utf-8', errors='ignore')
            execution.completed_at = timezone.now()
            
            if execution.started_at:
                execution.execution_time = (execution.completed_at - execution.started_at).total_seconds()
            
            # 根据退出码设置状态
            if process.returncode == 0:
                execution.status = 'completed'
                ExecutionLog.objects.create(
                    execution=execution,
                    level='info',
                    message="脚本执行成功完成"
                )
            else:
                execution.status = 'failed'
                ExecutionLog.objects.create(
                    execution=execution,
                    level='error',
                    message=f"脚本执行失败，退出码: {process.returncode}"
                )
            
            execution.save()
            
        except Exception as e:
            logger.error(f"执行Midscene脚本失败: {str(e)}")
            execution.status = 'failed'
            execution.stderr = str(e)
            execution.completed_at = timezone.now()
            execution.save()
            
            ExecutionLog.objects.create(
                execution=execution,
                level='error',
                message=f"执行异常: {str(e)}"
            )
            
            raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise
    
    async def _process_execution_result(self, execution: ScriptExecution, execution_dir: str):
        """处理执行结果"""
        try:
            # 查找报告文件
            report_dir = os.path.join(execution_dir, 'midscene_run', 'report')
            if os.path.exists(report_dir):
                # 设置报告路径
                execution.report_path = report_dir
                
                # 生成报告访问URL
                execution.report_url = f"/api/automation-scripts/executions/{execution.id}/report/"
                
                # 解析测试结果
                await self._parse_test_results(execution, report_dir)
            
            execution.save()
            
            # 记录结果处理日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message=f"结果处理完成，报告路径: {execution.report_path}"
            )
            
        except Exception as e:
            logger.error(f"处理执行结果失败: {str(e)}")
            ExecutionLog.objects.create(
                execution=execution,
                level='warning',
                message=f"结果处理失败: {str(e)}"
            )
    
    async def _parse_test_results(self, execution: ScriptExecution, report_dir: str):
        """解析测试结果"""
        try:
            # 查找结果JSON文件
            results_file = os.path.join(report_dir, 'results.json')
            if os.path.exists(results_file):
                with open(results_file, 'r', encoding='utf-8') as f:
                    results = json.load(f)
                
                # 解析统计信息
                execution.total_tests = results.get('total', 0)
                execution.passed_tests = results.get('passed', 0)
                execution.failed_tests = results.get('failed', 0)
            
        except Exception as e:
            logger.warning(f"解析测试结果失败: {str(e)}")


class AutomationScriptService:
    """自动化脚本服务"""
    
    @staticmethod
    def create_script(data: Dict, creator, project) -> AutomationScript:
        """创建脚本"""
        try:
            script = AutomationScript.objects.create(
                project=project,
                creator=creator,
                name=data['name'],
                description=data.get('description', ''),
                script_type=data['script_type'],
                test_cases_content=data['test_cases_content'],
                target_url=data.get('target_url', ''),
                viewport_width=data.get('viewport_width', 1280),
                viewport_height=data.get('viewport_height', 960),
                ai_model=data.get('ai_model', 'qwen-turbo'),
                api_key=data.get('api_key', ''),
                api_endpoint=data.get('api_endpoint', ''),
                execution_timeout=data.get('execution_timeout', 300),
                retry_count=data.get('retry_count', 1)
            )
            
            return script
            
        except Exception as e:
            logger.error(f"创建脚本失败: {str(e)}")
            raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_yaml_script(script: AutomationScript) -> AutomationScript:
        """生成YAML脚本"""
        try:
            if not script.api_key or not script.api_endpoint:
                raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise Exception("AI配置不完整，请检查API密钥和端点")
            
            # 初始化生成器
            generator = YAMLScriptGenerator(
                api_key=script.api_key,
                api_endpoint=script.api_endpoint,
                model=script.ai_model
            )
            
            # 生成YAML脚本
            yaml_content = generator.generate_yaml_script(
                test_cases=script.test_cases_content,
                script_type=script.script_type,
                target_url=script.target_url
            )
            
            # 更新脚本
            script.yaml_content = yaml_content
            script.status = 'generated'
            script.generated_at = timezone.now()
            script.save()
            
            return script
            
        except Exception as e:
            logger.error(f"生成YAML脚本失败: {str(e)}")
            script.status = 'failed'
            script.save()
            raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise
    
    @staticmethod
    async def execute_script(script: AutomationScript, executor_user) -> ScriptExecution:
        """执行脚本"""
        try:
            if not script.yaml_content:
                raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise Exception("脚本未生成，请先生成YAML脚本")
            
            if script.status not in ['generated', 'ready']:
                raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise Exception(f"脚本状态不正确: {script.status}")
            
            # 更新脚本状态
            script.status = 'running'
            script.save()
            
            # 初始化执行器
            executor = MidsceneExecutor()
            
            # 执行脚本
            execution = await executor.execute_script(script, executor_user)
            
            # 更新脚本状态
            if execution.status == 'completed':
                script.status = 'completed'
            else:
                script.status = 'failed'
            script.save()
            
            return execution
            
        except Exception as e:
            logger.error(f"执行脚本失败: {str(e)}")
            script.status = 'failed'
            script.save()
            raise
    
    @staticmethod
    def stop_execution(execution: ScriptExecution):
        """停止执行"""
        try:
            execution.status = 'cancelled'
            execution.completed_at = timezone.now()
            execution.save()
            
            # 如果有正在运行的进程，尝试终止
            if execution.process_id:
                try:
                    import psutil
                    process = psutil.Process(execution.process_id)
                    process.terminate()
                except:
                    pass
            
            # 记录日志
            ExecutionLog.objects.create(
                execution=execution,
                level='info',
                message='执行已被用户停止',
                step_name='系统操作'
            )
            
        except Exception as e:
            logger.error(f"停止执行失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_screenshots(execution: ScriptExecution) -> List[Dict]:
        """获取执行截图"""
        try:
            screenshots = []
            
            # 获取报告目录
            report_dir = os.path.join(settings.MEDIA_ROOT, 'automation_reports', str(execution.id))
            
            if os.path.exists(report_dir):
                # 查找截图文件
                for root, dirs, files in os.walk(report_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                            file_path = os.path.join(root, file)
                            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                            
                            screenshots.append({
                                'step': len(screenshots) + 1,
                                'url': f'/media/{relative_path.replace(os.sep, "/")}',
                                'description': file.replace('.png', '').replace('_', ' '),
                                'timestamp': timezone.now().isoformat()
                            })
            
            return screenshots
            
        except Exception as e:
            logger.error(f"获取截图失败: {str(e)}")
            return []
    
    @staticmethod
    def rerun_script(execution: ScriptExecution, user) -> ScriptExecution:
        """重新执行脚本"""
        try:
            # 创建新的执行记录
            new_execution = ScriptExecution.objects.create(
                script=execution.script,
                executor=user,
                status='pending'
            )
            
            # 异步执行脚本
            thread = threading.Thread(
                target=AutomationScriptService.execute_script,
                args=(new_execution,)
            )
            thread.daemon = True
            thread.start()
            
            return new_execution
            
        except Exception as e:
            logger.error(f"重新执行失败: {str(e)}")
            raise
    
    @staticmethod
    def generate_share_url(execution: ScriptExecution) -> str:
        """生成分享链接"""
        try:
            # 生成分享token
            import uuid
            share_token = str(uuid.uuid4())
            
            # 可以存储到缓存或数据库中
            # 这里简化处理，直接返回URL
            base_url = settings.FRONTEND_URL or 'http://localhost:8080'
            share_url = f"{base_url}/shared-report/{execution.id}?token={share_token}"
            
            return share_url
            
        except Exception as e:
            logger.error(f"生成分享链接失败: {str(e)}")
            raise
    
    @staticmethod
    def get_execution_stats(project_id=None, date_range=None, user=None) -> Dict:
        """获取执行统计"""
        try:
            queryset = ScriptExecution.objects.all()
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 项目过滤
            if project_id:
                queryset = queryset.filter(script__project_id=project_id)
            
            # 日期过滤
            if date_range and len(date_range) == 2:
                queryset = queryset.filter(
                    created_at__date__range=[date_range[0], date_range[1]]
                )
            
            # 统计数据
            total_executions = queryset.count()
            completed_executions = queryset.filter(status='completed').count()
            success_rate = (completed_executions / total_executions * 100) if total_executions > 0 else 0
            
            # 平均耗时
            completed_with_time = queryset.filter(
                status='completed',
                started_at__isnull=False,
                completed_at__isnull=False
            )
            
            avg_duration = 0
            if completed_with_time.exists():
                durations = []
                for execution in completed_with_time:
                    duration = (execution.completed_at - execution.started_at).total_seconds()
                    durations.append(duration)
                avg_duration = sum(durations) / len(durations)
            
            # 今日执行数
            today = timezone.now().date()
            today_executions = queryset.filter(created_at__date=today).count()
            
            return {
                'total_executions': total_executions,
                'success_rate': round(success_rate, 2),
                'avg_duration': round(avg_duration, 2),
                'today_executions': today_executions
            }
            
        except Exception as e:
            logger.error(f"获取统计数据失败: {str(e)}")
            raise
    
    @staticmethod
    def export_execution_reports(filters: Dict, user) -> bytes:
        """导出执行报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "执行报告"
            
            # 设置表头
            headers = [
                '脚本名称', '执行状态', '成功步骤', '失败步骤', '总步骤',
                '执行时长', '开始时间', '结束时间', '执行者', '项目'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 获取数据
            queryset = ScriptExecution.objects.select_related('script', 'executor', 'script__project')
            
            # 权限过滤
            if user and hasattr(user, 'project_permissions'):
                accessible_projects = user.project_permissions.values_list('project_id', flat=True)
                queryset = queryset.filter(script__project_id__in=accessible_projects)
            
            # 应用筛选条件
            if filters.get('script_name'):
                queryset = queryset.filter(script__name__icontains=filters['script_name'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('project_id'):
                queryset = queryset.filter(script__project_id=filters['project_id'])
            if filters.get('date_start') and filters.get('date_end'):
                queryset = queryset.filter(
                    created_at__date__range=[filters['date_start'], filters['date_end']]
                )
            
            # 填充数据
            for row, execution in enumerate(queryset.order_by('-created_at'), 2):
                duration = ''
                if execution.started_at and execution.completed_at:
                    duration = str(execution.completed_at - execution.started_at)
                
                ws.cell(row=row, column=1, value=execution.script.name)
                ws.cell(row=row, column=2, value=execution.get_status_display())
                ws.cell(row=row, column=3, value=execution.passed_tests or 0)
                ws.cell(row=row, column=4, value=execution.failed_tests or 0)
                ws.cell(row=row, column=5, value=execution.total_tests or 0)
                ws.cell(row=row, column=6, value=duration)
                ws.cell(row=row, column=7, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
                ws.cell(row=row, column=8, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
                ws.cell(row=row, column=9, value=execution.executor.username if execution.executor else '')
                ws.cell(row=row, column=10, value=execution.script.project.name if execution.script.project else '')
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到内存
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"导出报告失败: {str(e)}")
            raise
    
    @staticmethod
    def compare_executions(base_execution: ScriptExecution, compare_execution: ScriptExecution) -> Dict:
        """对比执行结果"""
        try:
            # 基础对比数据
            base_success_rate = (base_execution.passed_tests / base_execution.total_tests * 100) if base_execution.total_tests > 0 else 0
            compare_success_rate = (compare_execution.passed_tests / compare_execution.total_tests * 100) if compare_execution.total_tests > 0 else 0
            
            base_duration = 0
            compare_duration = 0
            
            if base_execution.started_at and base_execution.completed_at:
                base_duration = (base_execution.completed_at - base_execution.started_at).total_seconds()
            
            if compare_execution.started_at and compare_execution.completed_at:
                compare_duration = (compare_execution.completed_at - compare_execution.started_at).total_seconds()
            
            # 计算差异
            differences = {
                'success_rate_diff': round(compare_success_rate - base_success_rate, 2),
                'duration_diff': round(compare_duration - base_duration, 2),
                'steps_diff': (compare_execution.total_tests or 0) - (base_execution.total_tests or 0)
            }
            
            # 步骤对比（简化版本）
            step_comparison = []
            max_steps = max(base_execution.total_tests or 0, compare_execution.total_tests or 0)
            
            for i in range(1, max_steps + 1):
                step_comparison.append({
                    'step_number': i,
                    'step_name': f'步骤 {i}',
                    'description': f'测试步骤 {i}',
                    'base_status': 'success' if i <= (base_execution.passed_tests or 0) else 'failed',
                    'compare_status': 'success' if i <= (compare_execution.passed_tests or 0) else 'failed',
                    'base_duration': 1000 + i * 100,  # 模拟数据
                    'compare_duration': 1000 + i * 120,  # 模拟数据
                    'status_changed': False,
                    'status_improved': False,
                    'time_diff': 20,  # 模拟数据
                    'time_improved': False
                })
            
            return {
                'base_report': {
                    'id': base_execution.id,
                    'script_name': base_execution.script.name,
                    'success_steps': base_execution.passed_tests or 0,
                    'failed_steps': base_execution.failed_tests or 0,
                    'total_steps': base_execution.total_tests or 0,
                    'duration': base_duration
                },
                'compare_report': {
                    'id': compare_execution.id,
                    'script_name': compare_execution.script.name,
                    'success_steps': compare_execution.passed_tests or 0,
                    'failed_steps': compare_execution.failed_tests or 0,
                    'total_steps': compare_execution.total_tests or 0,
                    'duration': compare_duration
                },
                'differences': differences,
                'step_comparison': step_comparison
            }
            
        except Exception as e:
            logger.error(f"对比执行结果失败: {str(e)}")
            raise